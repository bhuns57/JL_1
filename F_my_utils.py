############## DEF FUNCTIONS SAVED AS A PY MODULE ###################
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def load_df_with_formulas(file_name="xl_tbl.xlsx", sheet_name="xl_tbl"):
    """
    Reads an Excel sheet, preserving formulas and structure.
    Returns a pandas DataFrame with formulas as strings.
    """
    wb = load_workbook(file_name, data_only=False)
    ws = wb[sheet_name]

    data = []
    for row in ws.iter_rows(values_only=False):
        row_data = []
        for cell in row:
            if cell.value is not None:
                if cell.data_type == 'f':
                    row_data.append(f"={cell.value}")
                else:
                    row_data.append(cell.value)
            else:
                row_data.append(None)
        data.append(row_data)

    headers = data[0]
    df = pd.DataFrame(data[1:], columns=headers)
    return df

def write_df_with_formulas(df, file_name="xl_tbl.xlsx", sheet_name="xl_tbl", table_name="xl_tbl"):
    """
    Writes a DataFrame to Excel, preserving formula strings.
    Rebuilds the target sheet and formats it as a table.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(value, str) and value.startswith("="):
                cell.value = value[1:]  # Strip '=' for Excel formula
                cell.data_type = 'f'
            else:
                cell.value = value

    # Create table
    nrows, ncols = df.shape
    table_range = f"A1:{get_column_letter(ncols)}{nrows+1}"
    table = Table(displayName=table_name, ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(file_name)




